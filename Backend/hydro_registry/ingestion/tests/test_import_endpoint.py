"""Канонический тестовый шов фичи (ADR/PRD): POST файла в /api/import/.

Мокается ТОЛЬКО LLM-порт; парсинг, раскладка и запись в PostGIS идут по-настоящему.
Требует среды с GDAL/GEOS и тестовой БД PostGIS (см. CLAUDE.md → Platform notes).
"""

import io
from unittest.mock import patch

from django.contrib.gis.geos import Point
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase
from openpyxl import Workbook

from ingestion.domain.types import MappingResult
from infrastructure.models import Canal, Sluice


def _canal_xlsx_upload():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "каналы"
    worksheet.append(["Наименование", "Водоисточник", "Год", "Пропускная способность", "Широта", "Долгота"])
    worksheet.append(["Канал А", "р. Иртыш", 1973, 3.0, 50.1, 80.2])
    worksheet.append(["Канал Б", "р. Иртыш", 1945, 1.5, 50.3, 80.4])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return SimpleUploadedFile(
        "kanaly.xlsx",
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


_CANAL_MAPPING = {0: "name", 1: "water_source", 2: "year_built", 3: "capacity", 4: "latitude", 5: "longitude"}


class _StubMapper:
    def map(self, *, facility_hint, columns):
        return MappingResult(facility_type="canal", mapping=dict(_CANAL_MAPPING))


def _two_sheet_xlsx_upload(correction_capacity=3.0):
    """Два листа с РАЗНЫМ порядком колонок, описывающие частично те же каналы."""
    workbook = Workbook()
    list1 = workbook.active
    list1.title = "Лист1"
    list1.append(["Наименование", "Водоисточник", "Год", "Расход", "Широта", "Долгота"])
    list1.append(["Канал А", "р. Иртыш", 1973, 3.0, 50.1, 80.2])
    list1.append(["Канал Б", "р. Иртыш", 1945, 1.5, 50.3, 80.4])
    correction = workbook.create_sheet("Корректировка")
    correction.append(["Водоисточник", "Наименование", "Год", "Расход", "Широта", "Долгота"])
    correction.append(["р. Иртыш", "Канал А", 1973, correction_capacity, 50.1, 80.2])
    correction.append(["р. Иртыш", "Канал В", 1928, 2.0, 51.0, 81.0])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return SimpleUploadedFile("multi.xlsx", buffer.read())


def _sluice_xlsx_upload():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "шлюзы"
    worksheet.append(["Наименование", "Водоисточник", "Год", "Затворы", "Привод"])
    worksheet.append(["Шлюз 1", "р. Иртыш", 1980, 5, "электрический"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return SimpleUploadedFile("sluice.xlsx", buffer.read())


class _SluiceStubMapper:
    def map(self, *, facility_hint, columns):
        return MappingResult(
            "sluice",
            {0: "name", 1: "water_source", 2: "year_built", 3: "gates_count", 4: "drive_type"},
        )


class _TwoSheetStubMapper:
    def map(self, *, facility_hint, columns):
        if facility_hint == "Корректировка":
            return MappingResult(
                "canal",
                {0: "water_source", 1: "name", 2: "year_built", 3: "capacity", 4: "latitude", 5: "longitude"},
            )
        return MappingResult("canal", dict(_CANAL_MAPPING))


class ImportEndpointTests(TestCase):
    @patch("ingestion.interfaces.views.resolve_schema_mapper", return_value=_StubMapper())
    def test_imports_canals_from_xlsx(self, _mapper):
        response = Client().post("/api/import/", {"file": _canal_xlsx_upload()})

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["created"], 2)
        self.assertEqual(Canal.objects.count(), 2)

        canal = Canal.objects.get(name="Канал А")
        self.assertEqual(canal.facility_type, "canal")
        self.assertEqual(canal.water_source, "р. Иртыш")
        self.assertEqual(canal.year_built, 1973)
        self.assertEqual(canal.capacity, 3.0)
        self.assertIsNotNone(canal.location)
        self.assertAlmostEqual(canal.location.x, 80.2)  # долгота
        self.assertAlmostEqual(canal.location.y, 50.1)  # широта

    @patch("ingestion.interfaces.views.resolve_schema_mapper", return_value=_StubMapper())
    def test_full_duplicate_is_skipped(self, _mapper):
        Canal.objects.create(
            facility_type="canal",
            name="Канал А",
            water_source="р. Иртыш",
            year_built=1973,
            capacity=3.0,
            location=Point(80.2, 50.1, srid=4326),
        )

        response = Client().post("/api/import/", {"file": _canal_xlsx_upload()})

        body = response.json()
        self.assertEqual(body["created"], 1)  # только Канал Б
        self.assertEqual(body["skipped_duplicates"], 1)
        self.assertEqual(Canal.objects.count(), 2)

    @patch("ingestion.interfaces.views.resolve_schema_mapper", return_value=_StubMapper())
    def test_value_divergence_reported_as_conflict(self, _mapper):
        Canal.objects.create(
            facility_type="canal",
            name="Канал А",
            water_source="р. Иртыш",
            year_built=1973,
            capacity=9.9,  # расходится с входящим 3.0
            location=Point(80.2, 50.1, srid=4326),
        )

        response = Client().post("/api/import/", {"file": _canal_xlsx_upload()})

        body = response.json()
        self.assertEqual(body["created"], 1)  # Канал Б создан, Канал А — нет
        self.assertEqual(len(body["conflicts"]), 1)
        self.assertEqual(body["conflicts"][0]["field"], "capacity")
        self.assertFalse(Canal.objects.filter(name="Канал А", capacity=3.0).exists())

    @patch("ingestion.interfaces.views.resolve_schema_mapper", return_value=_TwoSheetStubMapper())
    def test_cross_sheet_duplicate_deduped(self, _mapper):
        response = Client().post("/api/import/", {"file": _two_sheet_xlsx_upload()})

        body = response.json()
        # А, Б (Лист1) + В (Корректировка); повтор А с Корректировки пропущен.
        self.assertEqual(body["created"], 3)
        self.assertEqual(body["skipped_duplicates"], 1)
        self.assertEqual(Canal.objects.count(), 3)

    @patch("ingestion.interfaces.views.resolve_schema_mapper", return_value=_TwoSheetStubMapper())
    def test_cross_sheet_conflict_reported(self, _mapper):
        response = Client().post(
            "/api/import/", {"file": _two_sheet_xlsx_upload(correction_capacity=9.9)}
        )

        body = response.json()
        self.assertEqual(body["created"], 3)  # А, Б, В — А не пересоздан
        self.assertEqual(len(body["conflicts"]), 1)
        self.assertEqual(body["conflicts"][0]["sheet"], "Корректировка")
        self.assertEqual(Canal.objects.filter(name="Канал А").count(), 1)

    @patch("ingestion.interfaces.views.resolve_schema_mapper", return_value=_SluiceStubMapper())
    def test_imports_non_canal_type(self, _mapper):
        response = Client().post("/api/import/", {"file": _sluice_xlsx_upload()})

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["created"], 1)
        sluice = Sluice.objects.get(name="Шлюз 1")
        self.assertEqual(sluice.facility_type, "sluice")
        self.assertEqual(sluice.gates_count, 5)
        self.assertEqual(sluice.drive_type, "электрический")

    def test_missing_file_returns_400(self):
        response = Client().post("/api/import/", {})
        self.assertEqual(response.status_code, 400)

    def test_unsupported_extension_returns_400(self):
        upload = SimpleUploadedFile("dump.sql", b"INSERT INTO x VALUES (1);")
        response = Client().post("/api/import/", {"file": upload})
        self.assertEqual(response.status_code, 400)
