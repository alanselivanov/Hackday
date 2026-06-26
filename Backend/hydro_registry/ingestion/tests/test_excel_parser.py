"""Чистые тесты парсера Excel — без Django/GDAL, запускаются обычным unittest."""

import io
import os
import unittest

from openpyxl import Workbook

from ingestion.infrastructure.parsers.excel_parser import ExcelParser

_REAL_XLS = os.path.join(
    os.path.dirname(__file__), "..", "..", "..", "..", "датасет(1).xls"
)


def _make_xlsx(header, rows, sheet_title="каналы"):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title
    worksheet.append(header)
    for row in rows:
        worksheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


class ExcelParserTests(unittest.TestCase):
    def test_parses_flat_header_into_columns_and_rows(self):
        xlsx = _make_xlsx(
            ["Наименование", "Водоисточник", "Год"],
            [["Канал А", "р. Иртыш", 1973], ["Канал Б", "р. Иртыш", 1945]],
        )

        sheets = ExcelParser().parse(xlsx)

        self.assertEqual(len(sheets), 1)
        sheet = sheets[0]
        self.assertEqual(sheet.name, "каналы")
        self.assertEqual([c.name for c in sheet.columns], ["Наименование", "Водоисточник", "Год"])
        self.assertEqual(sheet.rows[0], ["Канал А", "р. Иртыш", 1973])

    def test_collects_samples_under_each_column(self):
        xlsx = _make_xlsx(
            ["Наименование", "Год"],
            [["Канал А", 1973], ["Канал Б", 1945], ["Канал В", 1928], ["Канал Г", 1955]],
        )

        sheet = ExcelParser().parse(xlsx)[0]

        year_column = sheet.columns[1]
        self.assertEqual(year_column.name, "Год")
        self.assertEqual(year_column.samples, [1973, 1945, 1928])  # максимум 3

    def test_blank_header_cells_are_skipped(self):
        xlsx = _make_xlsx(
            ["Наименование", None, "Год"],
            [["Канал А", "мусор", 1973]],
        )

        sheet = ExcelParser().parse(xlsx)[0]

        self.assertEqual([c.name for c in sheet.columns], ["Наименование", "Год"])
        self.assertEqual([c.index for c in sheet.columns], [0, 2])

    def test_all_numeric_first_data_row_is_not_dropped(self):
        # Регрессия: числовая первая строка данных не должна приниматься за
        # строку-нумерацию колонок (ряд нумерации начинается строго с 1).
        xlsx = _make_xlsx(
            ["A", "B", "C", "D"],
            [[10, 20, 30, 40], [11, 21, 31, 41]],
        )

        sheet = ExcelParser().parse(xlsx)[0]

        self.assertEqual(len(sheet.rows), 2)
        self.assertEqual(sheet.rows[0], [10, 20, 30, 40])


def _make_dirty_xlsx():
    """Госфайл-образец: full-width титры/группы, многоуровневая merge-шапка,
    строка-нумерация колонок, затем данные."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "каналы"
    # Баннеры-титры (full-width merge).
    worksheet.merge_cells("A1:E1")
    worksheet["A1"] = "Таблица 3"
    worksheet.merge_cells("A2:E2")
    worksheet["A2"] = "Технические характеристики всех каналов"
    # Многоуровневая шапка: вертикальные merge для простых колонок, горизонтальный для группы.
    worksheet["A3"] = "Наименование"
    worksheet.merge_cells("A3:A5")
    worksheet["B3"] = "Водоисточник"
    worksheet.merge_cells("B3:B5")
    worksheet["C3"] = "Год"
    worksheet.merge_cells("C3:C5")
    worksheet["D3"] = "Параметры"
    worksheet.merge_cells("D3:E3")
    worksheet["D4"] = "Ширина, м"
    worksheet["E4"] = "Глубина, м"
    # Строка-нумерация колонок.
    for col, num in zip("ABCDE", range(1, 6)):
        worksheet[f"{col}6"] = num
    # Баннер-группа внутри данных.
    worksheet.merge_cells("A7:E7")
    worksheet["A7"] = "Группа объектов 1"
    # Данные.
    worksheet["A8"], worksheet["B8"], worksheet["C8"], worksheet["D8"], worksheet["E8"] = (
        "Канал А", "р. Иртыш", 1973, 5.0, 2.0,
    )
    worksheet["A9"], worksheet["B9"], worksheet["C9"], worksheet["D9"], worksheet["E9"] = (
        "Канал Б", "р. Иртыш", 1945, 6.0, 3.0,
    )
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


class DirtyHeaderTests(unittest.TestCase):
    def test_builds_flat_names_from_multilevel_header(self):
        sheet = ExcelParser().parse(_make_dirty_xlsx())[0]

        names = [c.name for c in sheet.columns]
        self.assertEqual(
            names,
            ["Наименование", "Водоисточник", "Год", "Параметры / Ширина, м", "Параметры / Глубина, м"],
        )

    def test_drops_banner_and_number_rows_keeps_only_data(self):
        sheet = ExcelParser().parse(_make_dirty_xlsx())[0]

        self.assertEqual(len(sheet.rows), 2)
        first_cells = [row[0] for row in sheet.rows]
        self.assertEqual(first_cells, ["Канал А", "Канал Б"])
        # Ни титры, ни «Группа объектов», ни строка-нумерация не просочились.
        for row in sheet.rows:
            self.assertNotIn("Группа объектов 1", row)
            self.assertNotIn("Таблица 3", row)

    def test_samples_taken_from_data_not_header(self):
        sheet = ExcelParser().parse(_make_dirty_xlsx())[0]

        year_column = next(c for c in sheet.columns if c.name == "Год")
        self.assertEqual(year_column.samples, [1973, 1945])

    def test_non_sequential_number_row_is_detected(self):
        # Ряд нумерации с пропусками (как на листе «каналы»: 1,3,4,5) — всё равно
        # распознаётся как граница шапки/данных.
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "каналы"
        worksheet.merge_cells("A1:D1")
        worksheet["A1"] = "Таблица 3"
        worksheet["A2"] = "Имя"
        worksheet.merge_cells("A2:A3")
        worksheet["B2"] = "Группа"
        worksheet.merge_cells("B2:D2")
        worksheet["B3"], worksheet["C3"], worksheet["D3"] = "x", "y", "z"
        worksheet["A4"], worksheet["B4"], worksheet["C4"], worksheet["D4"] = 1, 3, 4, 5
        worksheet["A5"], worksheet["B5"], worksheet["C5"], worksheet["D5"] = "Канал", 10, 20, 30
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        sheet = ExcelParser().parse(buffer)[0]

        self.assertEqual([c.name for c in sheet.columns], ["Имя", "Группа / x", "Группа / y", "Группа / z"])
        self.assertEqual(sheet.rows, [["Канал", 10, 20, 30]])


@unittest.skipUnless(os.path.exists(_REAL_XLS), "реальный датасет(1).xls недоступен")
class RealXlsTests(unittest.TestCase):
    def _list1(self):
        with open(_REAL_XLS, "rb") as handle:
            sheets = ExcelParser().parse(handle)
        return next(s for s in sheets if s.name == "Лист1")

    def test_real_xls_header_and_junk(self):
        sheet = self._list1()
        names = [c.name for c in sheet.columns]

        self.assertIn("Наименование каналов", names)
        self.assertIn("Водоисточник", names)
        # Иерархия «из них → землян./облицов.» собрана во флэт-имя.
        self.assertTrue(any(" / " in n for n in names), names)
        # Заголовки таблицы и группы не попали в имена колонок.
        self.assertFalse(any("Таблица" in n or "Группа объектов" in n for n in names))
        # Данных много, и это реальные строки (есть водоисточник-текст).
        self.assertGreater(len(sheet.rows), 300)


if __name__ == "__main__":
    unittest.main()
