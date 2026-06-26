"""Чтение Excel в единый сырой вид: сетка ячеек + merge-диапазоны.

Отделяет формат/IO (.xls через xlrd, .xlsx через openpyxl) от логики разбора шапки.
Формат определяется по сигнатуре файла, а не по расширению.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any, BinaryIO


@dataclass(frozen=True)
class RawSheet:
    name: str
    grid: list[list[Any]]  # прямоугольная сетка значений (None для пустых)
    merged: list[tuple[int, int, int, int]]  # (r0, r1, c0, c1) — 0-based, half-open


def _rectangular(grid: list[list[Any]]) -> list[list[Any]]:
    width = max((len(row) for row in grid), default=0)
    return [list(row) + [None] * (width - len(row)) for row in grid]


def _read_xlsx(data: bytes) -> list[RawSheet]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), data_only=True)
    sheets: list[RawSheet] = []
    for worksheet in workbook.worksheets:
        grid = [list(row) for row in worksheet.iter_rows(values_only=True)]
        merged = [
            (rng.min_row - 1, rng.max_row, rng.min_col - 1, rng.max_col)
            for rng in worksheet.merged_cells.ranges
        ]
        sheets.append(RawSheet(worksheet.title, _rectangular(grid), merged))
    workbook.close()
    return sheets


def _read_xls(data: bytes) -> list[RawSheet]:
    import xlrd

    workbook = xlrd.open_workbook(file_contents=data, formatting_info=True)
    sheets: list[RawSheet] = []
    for sheet in workbook.sheets():
        grid = [
            [sheet.cell_value(r, c) for c in range(sheet.ncols)]
            for r in range(sheet.nrows)
        ]
        merged = list(sheet.merged_cells)  # уже (rlo, rhi, clo, chi), half-open
        sheets.append(RawSheet(sheet.name, _rectangular(grid), merged))
    return sheets


def read_workbook(file: BinaryIO) -> list[RawSheet]:
    data = file.read()
    if data[:2] == b"PK":  # zip-контейнер → .xlsx
        return _read_xlsx(data)
    if data[:4] == b"\xd0\xcf\x11\xe0":  # OLE2 → .xls
        return _read_xls(data)
    raise ValueError("Неподдерживаемый формат Excel-файла.")
