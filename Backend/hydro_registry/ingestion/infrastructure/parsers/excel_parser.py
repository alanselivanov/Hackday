"""Парсер Excel. На срезе #01 — один лист с плоской шапкой (первая строка = заголовки).

Многоуровневая шапка, протягивание merge и отброс мусорных строк добавляются в #02.
"""

from __future__ import annotations

from typing import Any, BinaryIO

from openpyxl import load_workbook

from ...domain.types import ColumnSample, ParsedSheet

_MAX_SAMPLES = 3


def _samples_for(rows: list[list[Any]], col_index: int) -> list[Any]:
    samples: list[Any] = []
    for row in rows:
        if col_index >= len(row):
            continue
        value = row[col_index]
        if value is None:
            continue
        if isinstance(value, str) and value.strip() == "":
            continue
        samples.append(value)
        if len(samples) >= _MAX_SAMPLES:
            break
    return samples


class ExcelParser:
    def parse(self, file: BinaryIO) -> list[ParsedSheet]:
        workbook = load_workbook(file, read_only=True, data_only=True)
        sheets: list[ParsedSheet] = []
        for worksheet in workbook.worksheets:
            grid = [list(row) for row in worksheet.iter_rows(values_only=True)]
            if not grid:
                continue

            header, *data = grid  # элементы grid уже list, повторно не копируем
            rows = data
            columns = [
                ColumnSample(
                    index=i,
                    name=str(name).strip(),
                    samples=_samples_for(rows, i),
                )
                for i, name in enumerate(header)
                if name is not None and str(name).strip() != ""
            ]
            sheets.append(
                ParsedSheet(name=worksheet.title, columns=columns, rows=rows)
            )
        workbook.close()
        return sheets
